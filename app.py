import streamlit as st
from groq import Groq
import json
import re
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 페이지 설정 ───────────────────────────────────────────────
st.set_page_config(
    page_title="팀즈 보고서 자동화",
    page_icon="📋",
    layout="centered"
)

# ── 스타일 ────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@400;500;700&display=swap');

html, body, [class*="css"] {
    font-family: 'Noto Sans KR', sans-serif;
}

.main-title {
    font-size: 2rem;
    font-weight: 700;
    color: #1a1a2e;
    margin-bottom: 0.2rem;
}

.sub-title {
    color: #666;
    font-size: 0.95rem;
    margin-bottom: 2rem;
}

.step-badge {
    display: inline-block;
    background: #4f46e5;
    color: white;
    border-radius: 50%;
    width: 24px;
    height: 24px;
    text-align: center;
    line-height: 24px;
    font-size: 0.8rem;
    font-weight: 700;
    margin-right: 8px;
}

.info-box {
    background: #f0f4ff;
    border-left: 4px solid #4f46e5;
    padding: 12px 16px;
    border-radius: 0 8px 8px 0;
    font-size: 0.88rem;
    color: #444;
    margin-bottom: 1rem;
}

.success-box {
    background: #f0fdf4;
    border-left: 4px solid #22c55e;
    padding: 12px 16px;
    border-radius: 0 8px 8px 0;
    font-size: 0.9rem;
    color: #166534;
}

.stTextArea textarea {
    font-size: 0.88rem !important;
}
</style>
""", unsafe_allow_html=True)

# ── 헤더 ──────────────────────────────────────────────────────
st.markdown('<p class="main-title">📋 팀즈 대화 → 보고서 자동화</p>', unsafe_allow_html=True)
st.markdown('<p class="sub-title">팀즈 대화를 붙여넣으면 Excel 보고서를 자동으로 만들어 드려요</p>', unsafe_allow_html=True)

st.divider()

# ── STEP 1: API Key ───────────────────────────────────────────
st.markdown('<span class="step-badge">1</span> **Groq API Key 입력**', unsafe_allow_html=True)
st.markdown('<div class="info-box">🔑 <a href="https://console.groq.com" target="_blank">console.groq.com</a> 에서 발급받은 API Key를 입력하세요. 입력한 Key는 이 앱에 저장되지 않습니다.</div>', unsafe_allow_html=True)

api_key = st.text_input(
    "API Key",
    type="password",
    placeholder="gsk_...",
    label_visibility="collapsed"
)

st.divider()

# ── STEP 2: 대화 입력 ─────────────────────────────────────────
st.markdown('<span class="step-badge">2</span> **팀즈 대화 붙여넣기**', unsafe_allow_html=True)
st.markdown('<div class="info-box">💬 팀즈에서 대화를 복사해서 아래에 붙여넣으세요. 날짜, 참여자, 내용이 포함될수록 보고서가 더 정확해요.</div>', unsafe_allow_html=True)

chat_text = st.text_area(
    "대화 내용",
    height=280,
    placeholder="""예시:
2024-01-15 오전 10:00
김철수: 이번 프로젝트 일정 어떻게 됩니까?
이영희: 다음 주까지 기획서 완성하고, 월말까지 개발 완료 예정입니다.
박민준: 예산은 500만원으로 확정됐습니다.
김철수: 알겠습니다. 그럼 이영희님이 PM 맡아주세요.
이영희: 네, 진행하겠습니다.""",
    label_visibility="collapsed"
)

# ── Groq 분석 함수 ────────────────────────────────────────────
def analyze_chat(api_key: str, chat: str, _: str = "") -> dict:
    client = Groq(api_key=api_key)

    prompt = f"""아래 팀즈 대화를 분석해서 미팅 공유 보고서 형식의 JSON을 반환해줘.
반드시 아래 JSON 형식만 반환하고 다른 텍스트는 절대 포함하지 마.

{{
  "title": "미팅 주제 또는 프로젝트명 (예: 2024 마케팅 전략)",
  "meeting_datetime": "미팅 날짜와 시간 (예: 2024-01-15 오전 10:00, 없으면 대화 날짜)",
  "participants": ["참여자1", "참여자2"],
  "background": "미팅을 하게 된 배경 (1-2문장, 왜 이 미팅을 하게 됐는지)",
  "purpose": "미팅의 목적 (1-2문장, 이 미팅에서 무엇을 달성하려 했는지)",
  "content": "미팅에서 논의된 내용을 계층형 불릿으로 작성. 번호는 사용하지 말 것. 형식: 대분류 항목은 '• 항목명', 세부사항은 '  - 세부내용' 으로 들여쓰기. 예시:\n• 사업 인지하게 된 배경\n  - 국지호 차장의 사업 이관 건\n• 구매 예산\n  - 2,000만원 이내"
}}

팀즈 대화:
{chat}"""

    response = client.chat.completions.create(
        model="llama-3.3-70b-versatile",
        messages=[
            {"role": "system", "content": "You are a JSON-only assistant. Always respond with valid JSON and nothing else."},
            {"role": "user", "content": prompt}
        ],
        max_tokens=2000,
        response_format={"type": "json_object"}
    )

    raw = response.choices[0].message.content.strip()

    # JSON 블록만 추출 (코드펜스 등 섞여도 안전하게)
    match = re.search(r'\{.*\}', raw, re.DOTALL)
    if match:
        raw = match.group(0)
    return json.loads(raw)


# ── Excel 보고서 생성 함수 ────────────────────────────────────
def create_xlsx(data: dict) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "미팅공유"

    COLOR_PRIMARY = "4F46E5"
    COLOR_BG_ALT  = "F0F4FF"
    COLOR_WHITE   = "FFFFFF"
    COLOR_GRAY    = "888888"
    COLOR_DARK    = "1A1A2E"

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 60

    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    row = 1

    # ── 제목: OOO 미팅공유 ──
    ws.merge_cells(f"A{row}:B{row}")
    cell = ws[f"A{row}"]
    cell.value = f"{data.get('title', '미팅')} 미팅공유"
    cell.font = Font(name="맑은 고딕", size=15, bold=True, color=COLOR_WHITE)
    cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 34
    row += 2

    def calc_height(text: str) -> int:
        if not text:
            return 18
        lines = text.split("\n")
        total = sum(max(1, -(-len(line) // 55)) for line in lines)
        return max(18, total * 15)

    def write_field(label: str, value: str):
        nonlocal row
        label_cell = ws[f"A{row}"]
        label_cell.value = f"* {label} :"
        label_cell.font = Font(name="맑은 고딕", size=10, bold=True, color=COLOR_DARK)
        label_cell.fill = PatternFill("solid", fgColor=COLOR_BG_ALT)
        label_cell.alignment = Alignment(horizontal="left", vertical="top", indent=1)
        label_cell.border = border
        ws.row_dimensions[row].height = calc_height(value)

        value_cell = ws[f"B{row}"]
        value_cell.value = value
        value_cell.font = Font(name="맑은 고딕", size=10, color=COLOR_DARK)
        value_cell.fill = PatternFill("solid", fgColor=COLOR_WHITE)
        value_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
        value_cell.border = border
        row += 1

    participants = ", ".join(data.get("participants", []))

    write_field("미팅일시", data.get("meeting_datetime", datetime.now().strftime("%Y-%m-%d")))
    write_field("미팅인원", participants)
    write_field("미팅배경", data.get("background", ""))
    write_field("미팅목적", data.get("purpose", ""))
    write_field("미팅내용", data.get("content", ""))

    row += 1

    # ── 하단 ──
    ws.merge_cells(f"A{row}:B{row}")
    cell = ws[f"A{row}"]
    cell.value = f"자동 생성된 보고서  |  {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    cell.font = Font(name="맑은 고딕", size=8, color=COLOR_GRAY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── 생성 버튼 ─────────────────────────────────────────────────
st.divider()
col1, col2, col3 = st.columns([1, 2, 1])
with col2:
    generate_btn = st.button("🚀 보고서 생성하기", use_container_width=True, type="primary")

if generate_btn:
    if not api_key:
        st.error("❌ API Key를 입력해주세요.")
    elif not chat_text.strip():
        st.error("❌ 팀즈 대화를 입력해주세요.")
    elif len(chat_text.strip()) < 20:
        st.error("❌ 대화 내용이 너무 짧습니다. 더 많은 내용을 붙여넣어 주세요.")
    else:
        with st.spinner("🤖 대화를 분석하고 보고서를 작성 중입니다..."):
            try:
                report_data = analyze_chat(api_key, chat_text, "")
                xlsx_bytes = create_xlsx(report_data)
                st.session_state["report_data"] = report_data
                st.session_state["xlsx_bytes"] = xlsx_bytes
            except json.JSONDecodeError:
                st.error("❌ 분석 중 오류가 발생했습니다. 대화 내용을 더 구체적으로 입력해주세요.")
            except Exception as e:
                st.error(f"❌ 오류가 발생했습니다: {str(e)}")

# ── 결과 표시 (session_state 기반) ───────────────────────────
if "report_data" in st.session_state and "xlsx_bytes" in st.session_state:
    report_data = st.session_state["report_data"]
    xlsx_bytes = st.session_state["xlsx_bytes"]

    st.markdown('<div class="success-box">✅ 보고서가 완성됐습니다! 아래 버튼으로 다운로드하세요.</div>', unsafe_allow_html=True)
    st.markdown("")

    with st.expander("📄 보고서 미리보기", expanded=True):
        st.markdown(f"### {report_data.get('title', '')} 미팅공유")
        st.markdown(f"**\\* 미팅일시 :** {report_data.get('meeting_datetime', '')}")
        st.markdown(f"**\\* 미팅인원 :** {', '.join(report_data.get('participants', []))}")
        st.markdown(f"**\\* 미팅배경 :** {report_data.get('background', '')}")
        st.markdown(f"**\\* 미팅목적 :** {report_data.get('purpose', '')}")
        st.markdown("**\\* 미팅내용 :**")
        st.write(report_data.get("content", ""))

    filename = f"보고서_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    col_a, col_b, col_c = st.columns([1, 2, 1])
    with col_b:
        st.download_button(
            label="📥 Excel 보고서 다운로드",
            data=xlsx_bytes,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

# ── 하단 안내 ─────────────────────────────────────────────────
st.divider()
st.caption("🔒 입력한 API Key와 대화 내용은 서버에 저장되지 않습니다.")
